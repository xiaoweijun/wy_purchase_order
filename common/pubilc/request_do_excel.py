# -*- coding: utf-8 -*-
# @Author : xiaowj
# @Time : 2022/2/26 17:06

from openpyxl import load_workbook
from common.pubilc.request_getconfig import GetConfig
from common.pubilc import request_getpath
from common.pubilc.request_getdata import GetData
from common.pubilc import request_logging
from common.pubilc.request_do_regx import DoRegx


class GetTestData:
    def __init__(self):
        self.table_name = request_getpath.testdata_path
        print("table_name是：{0}".format(self.table_name))
    # 拿到需要登录的用户
    def get_login_user(self):
        test_data = []
        wb_data = load_workbook(self.table_name)
        # 2、定位表单
        sheet_data = wb_data["login"]  # 传表单名 sheet

        sheet_user = wb_data["user"]

        for j in range(2, sheet_user.max_row+1):
            data = sheet_data.cell(2, 3).value
            res = {}
            if data.find("${tel}") != -1:
                data = eval(data.replace("${tel}", str(sheet_user.cell(j,1).value)))
                res["data"] = data
                res["role_code"] = sheet_user.cell(j,2).value
                res["headers"] = sheet_data.cell(2,4).value
                res["method"] = sheet_data.cell(2,5).value
                res["assert_info"] = sheet_data.cell(2,6).value
                res["role_cookie"] = sheet_user.cell(j,4).value
            test_data.append(res)

        return test_data





    #从excel获取数据并存储到列表中
    def get_data(self,mode1):

        print("mode:",mode1)
        mode = eval(mode1)
        test_data = []
        # 1、打开excel
        for key in mode:
            wb = load_workbook(self.table_name)
            #print("key 是:{0},key的类型是:{1}",key,type(mode(key)))
            # 2、定位表单
            sheet = wb[key]  # 传表单名 sheet
            print("mode[key]>>>" ,mode[key])
            # 3、定位单元格，行、列值
            if mode[key] == "all":
                for j in range(2, sheet.max_row+1):
                    res = {}
                    res["case_id"] = sheet.cell(j, 1).value
                    res["url"] = str(getattr(GetData,"ip_adress")+sheet.cell(j, 2).value)
                    # print("ipadress >>>{0}".format(getattr(GetData,"ip_adress")))
                    # print("url>>>{0}".format(type(res["url"])))
                    # print("拿到的url >>>{0}".format(str(sheet.cell(j, 2).value)))
                    #res["data"] = sheet.cell(j, 4).value

                    new_data = DoRegx.do_regx(sheet.cell(j, 4).value)
                    res["data"] = new_data

                    res["headers"] = sheet.cell(j, 5).value
                    res["role_cookie"] = sheet.cell(j, 6).value
                    res["method"] = sheet.cell(j, 7).value
                    res["assert_info"] = sheet.cell(j, 8).value
                    res["sheet_name"] = key

                    test_data.append(res)

            else:
                for i in mode[key]:
                    res = {}
                    res["case_id"] = sheet.cell(i+1, 1).value
                    res["url"] = getattr(GetData,"ip_adress")+sheet.cell(i+1, 2).value
                    #res["data"] = sheet.cell(i+1, 4).value

                    new_data = DoRegx.do_regx(sheet.cell(i+1, 4).value)
                    res["data"] = new_data

                    res["headers"] = sheet.cell(i + 1, 5).value
                    res["role_cookie"] = sheet.cell(i+1, 6).value
                    res["method"] = sheet.cell(i+1, 7).value
                    res["assert_info"] = sheet.cell(i+1, 8).value
                    res["sheet_name"] = key

                    test_data.append(res)
        return test_data


    #写回测试结果
    def wirte_data(self,sheet_name,num,value,TestResult):
        # 1、打开excel
        wb = load_workbook(self.table_name)

        # 2、定位表单
        sheet = wb[sheet_name]  # 传表单名 sheet
        sheet.cell(num+1,9).value = value
        sheet.cell(num+1,10).value = TestResult
        wb.save(self.table_name)

    # 更新房号、用户名后缀
    def updata_roomno(self,sheet_name,value):
        # 1、打开excel
        wb = load_workbook(self.table_name)

        # 2、定位表单
        sheet = wb[sheet_name]  # 传表单名 sheet
        sheet.cell(1, 4).value = value
        wb.save(self.table_name)

    # 保存获取到的userid 和 cookier   暂时用不上，直接用反射替代写入、取值
    def wirte_userid(self,sheet_name,value,cookier,role_code):
        # 1、打开excel
        wb = load_workbook(self.table_name)

        # 2、定位表单
        sheet = wb[sheet_name]  # 传表单名 sheet
        if role_code=="zc":
            sheet.cell(2, 3).value = value
            sheet.cell(2, 5).value = cookier
        elif role_code=="zcjl":
            sheet.cell(3, 3).value = value
            sheet.cell(3, 5).value = cookier
        elif role_code=="cw":
            sheet.cell(4, 3).value = value
            sheet.cell(4, 5).value = cookier
        elif role_code=="jszy":
            sheet.cell(5, 3).value = value
            sheet.cell(5, 5).value = cookier
        elif role_code=="tz":
            sheet.cell(6, 3).value = value
            sheet.cell(6, 5).value = cookier
        elif role_code=="jszz":
            sheet.cell(7, 3).value = value
            sheet.cell(7, 5).value = cookier
        elif role_code=="qyz":
            sheet.cell(8, 3).value = value
            sheet.cell(8, 5).value = cookier
        elif role_code=="zjb":
            sheet.cell(9, 3).value = value
            sheet.cell(9, 5).value = cookier
        elif role_code=="xftz":
            sheet.cell(10, 3).value = value
            sheet.cell(10, 5).value = cookier

        elif role_code=="xfjszy":
            sheet.cell(11, 3).value = value
            sheet.cell(11, 5).value = cookier

        elif role_code=="xfjszz":
            sheet.cell(12, 3).value = value
            sheet.cell(12, 5).value = cookier
        elif role_code=="xfcw":
            sheet.cell(13, 3).value = value
            sheet.cell(13, 5).value = cookier


        wb.save(self.table_name)








if __name__ == '__main__':

    # root_dir = os.path.dirname(os.path.abspath('.'))  # 获取当前文件所在目录的上一级目录，即项目所在目录
    # print(root_dir)
    mode = GetConfig.get_purchase_mode_config()
    r= GetTestData().get_data(mode)
    #
    # print(len(r))
    # print(r)
    #r = GetTestData().get_login_user()
    print(r)