# -*- coding: utf-8 -*-
# @Author : xiaowj
# @Time : 2022/2/26 22:35
from openpyxl import load_workbook
from common.pubilc import request_getpath
from common.pubilc.request_do_sql import DoSql
class GetData:
    purchase_code = load_workbook(request_getpath.report_and_user_path)["report_data"].cell(12, 3).value  # 认购单号

    reportID = "@reportID@"  #报备单id
    tradeId = "@tradeId@"   #认购单id
    #tradeId = DoSql().do_sql("SELECT id from yy_purchase_order where purchase_code ='{0}';".format(purchase_code))
    ip_adress = "https://pre.yunjinji.cn"
    login_url = "https://pre.yunjinji.cn/user-center/auth/staff/login"  #登录的url
    # ip_adress = "https://release.yunjinji.cn"  # release环境
    # login_url = "https://release.yunjinji.cn/user-center/auth/staff/login"

    #驻场线提成方式
    standCutCaseType ="@standCutCaseType@"
    fixedAmount="@fixedAmount@"
    standManagerCutCaseType= "@standManagerCutCaseType@"
    standManagerFixedAmount= "@standManagerFixedAmount@"
    areaManagerCutCaseType = "@areaManagerCutCaseType@"
    areaManagerFixedAmount = "@areaManagerFixedAmount@"
    #拓展线提成方式
    expandCutCaseType = "@expandCutCaseType@"
    expandFixedAmount = "@expandFixedAmount@"
    expandMgrCutCaseType = "@expandMgrCutCaseType@"
    expandMgrFixedAmount = "@expandMgrFixedAmount@"
    expandDirectorCutCaseType = "@expandDirectorCutCaseType@"
    expandDirectorFixedAmount = "@expandDirectorFixedAmount@"

    #驻场提成信息
    zc_cut = "@zc_cut@"
    areaManager = "@areaManager@"
    areaManagerId = "@areaManagerId@"
    standManager = "@standManager@"
    standManagerId = "@standManagerId@"
    standManagerRatio = "@standManagerRatio@"
    expandRatios = "@expandRatios@"
    #第一拓展信息
    ex1_userId = "@ex1_userId@"
    ex1_userName = "@ex1_userName@"
    ex1_ratio = "@ex1_ratio@"
    ex1_superiorId = "@ex1_superiorId@"
    ex1_superiorName = "@ex1_superiorName@"
    ex1_superiorSuperiorId = "@ex1_superiorSuperiorId@"
    ex1_superiorSuperiorName = "@ex1_superiorSuperiorName@"
    #第二拓展信息
    ex2_userId = "@ex2_userId@"
    ex2_userName = "@ex2_userName@"
    ex2_ratio = "@ex2_ratio@"
    ex2_superiorId = "@ex2_superiorId@"
    ex2_superiorName = "@ex2_superiorName@"
    ex2_superiorSuperiorId = "@ex2_superiorSuperiorId@"
    ex2_superiorSuperiorName = "@ex2_superiorSuperiorName@"
    #第三拓展信息
    ex3_userId = "@ex3_userId@"
    ex3_userName = "@ex3_userName@"
    ex3_ratio = "@ex3_ratio@"
    ex3_superiorId = "@ex3_superiorId@"
    ex3_superiorName = "@ex3_superiorName@"
    ex3_superiorSuperiorId = "@ex3_superiorSuperiorId@"
    ex3_superiorSuperiorName = "@ex3_superiorSuperiorName@"

    #从excel中获取账号
    zc_tel= load_workbook(request_getpath.report_and_user_path)["user"].cell(2,1).value #驻场
    zcjl_tel= load_workbook(request_getpath.report_and_user_path)["user"].cell(3,1).value #驻场经理
    cw_tel = load_workbook(request_getpath.report_and_user_path)["user"].cell(4, 1).value #财务
    jszy_tel = load_workbook(request_getpath.report_and_user_path)["user"].cell(5, 1).value #结算专员
    tz_tel = load_workbook(request_getpath.report_and_user_path)["user"].cell(6, 1).value #拓展
    jszz_tel = load_workbook(request_getpath.report_and_user_path)["user"].cell(7, 1).value #结算组长
    qyz_tel = load_workbook(request_getpath.report_and_user_path)["user"].cell(8, 1).value #区域总
    zjb_tel = load_workbook(request_getpath.report_and_user_path)["user"].cell(9, 1).value #总经办
    xftz_tel = load_workbook(request_getpath.report_and_user_path)["user"].cell(10, 1).value #销方拓展

    # 从excel获取userid
    zc_userid = "@zc_userid@"
    zcjl_userid = "@zcjl_userid@"
    cw_userid = "@cw_userid@"
    jszy_userid = "@jszy_userid@"
    tz_userid = "@tz_userid@"
    jszz_userid = "@jszz_userid@"
    qyz_userid = "@qyz_userid@"
    zjb_userid = "@zjb_userid@"
    xftz_userid = "@xftz_userid@"
    xfjszy_userid = "@xfjszy_userid@"
    xfjszz_userid = "@xfjszz_userid@"
    xfcw_userid = "@xfcw_userid@"


    # 用户cookie
    zc_cookie = None
    zcjl_cookie = None
    cw_cookie = None
    jszy_cookie = None
    tz_cookie = None
    jszz_cookie = None
    qyz_cookie = None
    zjb_cookie = None
    xftz_cookie = None
    xfjszy_cookie = None
    xfjszz_cookie = None
    xfcw_cookie = None

    salePlatformId = load_workbook(request_getpath.report_and_user_path)["report_data"].cell(3, 3).value #平台id
    projectId= load_workbook(request_getpath.report_and_user_path)["report_data"].cell(4, 3).value #楼盘id
    reportTime = load_workbook(request_getpath.report_and_user_path)["report_data"].cell(5, 3).value #报备时间
    visitTime = load_workbook(request_getpath.report_and_user_path)["report_data"].cell(6, 3).value  #到访时间
    purchaseDate = load_workbook(request_getpath.report_and_user_path)["report_data"].cell(7, 3).value  #认购时间
    roomNo_num = load_workbook(request_getpath.report_and_user_path)["report_data"].cell(1, 4).value
    roomNo = load_workbook(request_getpath.report_and_user_path)["report_data"].cell(1, 3).value + str(roomNo_num)   #房号
    customName = load_workbook(request_getpath.report_and_user_path)["report_data"].cell(2, 3).value + str(roomNo_num)  # 客户姓名
    agent_salePlatformId = load_workbook(request_getpath.report_and_user_path)["report_data"].cell(8, 3).value #代理成交平台id
    deptId = load_workbook(request_getpath.report_and_user_path)["report_data"].cell(9, 3).value # 门店id
    agent_user_phone = load_workbook(request_getpath.report_and_user_path)["report_data"].cell(10, 3).value # 无忧经纪人手机号
    paymentMethod = load_workbook(request_getpath.report_and_user_path)["report_data"].cell(11, 3).value # 付款方式
    #purchase_code = load_workbook(request_getpath.report_and_user_path)["report_data"].cell(12, 3).value # 认购单号


    agent_user_id = DoSql().do_sql("SELECT userId from yy_user where account = '{0}' and isDel = 0 and status =1 and departmentId = {1}".format(agent_user_phone,deptId))


    # 结佣方案id
    pay_id = DoSql().do_sql("SELECT b.id from  yy_comm_project as a LEFT JOIN yy_pay_comm_plan as b on a.cycle_id = b.cycle_id where a.project_id = '{0}' and a.del = 0 and  '{1}' BETWEEN valid_start_time and valid_end_time ORDER BY b.id DESC LIMIT 1 ;".format(projectId,purchaseDate))
    #团购方案
    groupbuyPlanId = DoSql().do_sql("SELECT id FROM yy_comm_groupbuy_plan where cycle_id in (SELECT id FROM yy_comm_project_cycle where project_id = (SELECT id from yy_comm_project where project_id = {0})) and  '{1}' BETWEEN valid_start_time and valid_end_time and receivable_groupbuy != 0  LIMIT 1 ;".format(projectId,purchaseDate))

    #收佣方案id
    rec_id = DoSql().do_sql("SELECT b.id from  yy_comm_project as a LEFT JOIN yy_rec_comm_plan as b on a.cycle_id = b.cycle_id where a.project_id = '{0}' and a.del = 0 and  '{1}' BETWEEN valid_start_time and valid_end_time ORDER BY b.id DESC LIMIT 1;".format(projectId,purchaseDate))
    # 销方平台公司账户
    salePlatformAccountId = DoSql().do_sql("SELECT id from yy_comm_self_account where platform_id = {0} limit 1".format(agent_salePlatformId))
    # 平台公司id collectingCompanyId
    collectingCompanyId = DoSql().do_sql("SELECT company_id from yy_comm_project_cycle where project_id = (SELECT id FROM yy_comm_project where project_id = {0}) LIMIT 1;".format(projectId))
    # 开发商账户
    dev_account = DoSql().do_sql("SELECT id from yy_comm_dev_account where project_id = {0} limit 1".format(projectId))
    # 门店账户
    dept_account = DoSql().do_sql("SELECT baid from yy_department_ba where departmentId = {0} and isDel = 0 and state = 1 limit 1".format(str(deptId)))
    # rec_ticketId 收佣单id
    rec_ticketId = "@rec_ticketId@"
    # pay_ticket 结佣单id
    pay_ticketId = "@pay_ticketId@"
    #认购单的应收佣金
    receivable_commision = "@receivable_commision@"
    # 认购单的应结佣金
    payable_commission = "@payable_commission@"
    #公司账户id、名称
    account_id = DoSql().do_sql("SELECT id FROM yy_comm_self_account where platform_id = '{0}' LIMIT 1;".format(salePlatformId))
    account_name = DoSql().do_sql("SELECT account_name FROM yy_comm_self_account where platform_id = '{0}' LIMIT 1;".format(salePlatformId))

if __name__ == '__main__':
    setattr(GetData,"Cookie","123456")   #设置属性值
    print("判断是否有这个属性：",hasattr(GetData, "Cookie"))   #判断是否有这个属性
    print(getattr(GetData,"Cookie"))  #获取属性值
    delattr(GetData,"Test_attribute")  # 删除这个属性
    print("判断是否有这个属性：", hasattr(GetData, "Test_attribute"))
    print(getattr(GetData,"zc_tel1"))

